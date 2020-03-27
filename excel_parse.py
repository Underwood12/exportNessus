import csv
from openpyxl import Workbook
import pandas as pd
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog


rt = tk.Tk()
rt.withdraw()


def csvs_to_excel(csvs):
    wb = Workbook()
    csv.register_dialect('colons', delimiter=';')
    dest_filename = file_save()
    print(dest_filename)
    for file in csvs:
        tmp_ws = wb.create_sheet(file,csvs.index(file))
        with open(file) as f:
            reader = csv.reader(f,delimiter=';')
            for row in reader:
                tmp_ws.append(row)
        tmp_ws.auto_filter.ref = "A1:" + chr(tmp_ws.max_column + 64) + str(tmp_ws.max_row)
    wb.save(filename = dest_filename)
   #redo all worksheets in pivot mode 
    
    for file in csvs:
        with pd.ExcelWriter(dest_filename, engine="openpyxl") as writer:
            writer.book = wb
            df1 = pd.DataFrame(pd.read_csv(file, sep=";"))
             
            df = pd.pivot_table(df1, index="Plugin Name",values="Host IP", aggfunc="count")
            print(df)
            df.to_excel(writer, "piv_" + file, index=True)
    

def file_save():
    return filedialog.asksaveasfilename(defaultextension=".xlsx")

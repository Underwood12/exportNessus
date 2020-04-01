import openpyxl
import ipaddress
import report_host, report_item

#this is the risk matrix as a dictionary
# 1. By using the cvss score select the right dictionary.
# 2. Using the bus_crit from the host select the right sub list.
# 3. If an exploit is available according to tenable, select the right value if not select left.

DICT_L3_CVSS  =  [[1,2] , [5,6]   , [8,9]]
DICT_3T6_CVSS =  [[3,4] , [7,8]   , [10,11]]
DICT_7T8_CVSS =  [[5,6] , [9,10]  , [12,13]]
DICT_H8_CVSS  =  [[7,8] , [11,12] , [14,15]]




ITOP_COLUMNS = ["IP", "FQDN", "Status", "Business criticality", "Name"]
crit_dict = {}
row = 0

path = "itop_hosts.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active
row = ws.max_row

def init_dict():
    for i in range(1, row +1):
        ip      = ws.cell(row=i, column=ITOP_COLUMNS.index("IP") + 1).value
        fqdn    = ws.cell(row=i, column=ITOP_COLUMNS.index("FQDN") + 1).value
        stat    = ws.cell(row=i, column=ITOP_COLUMNS.index("Status") +1).value
        crit    = ws.cell(row=i, column=ITOP_COLUMNS.index("Business criticality") +1).value
        crit_dict[ip] = crit
        #print('{}:{}:{} --> {}'.format(ip, fqdn, stat, crit))

def set_access_score(creds, ip_is_private):
    score = 1
    if not ip_is_private: score += 2
    if not creds : score += 2
    return score

def get_crit_from(ip):
    try: return crit_dict[ip]
    except: return 'High by Default'

def is_ip_private(ip):
    return ipaddress.ip_address(ip).is_private

def get_cvss_list(cvss):
    if float(cvss) < 3: return DICT_L3_CVSS
    elif 3 <= float(cvss) <= 6 : return DICT_3T6_CVSS
    elif 6 <= float(cvss) <= 8 : return DICT_7T8_CVSS
    elif float(cvss) > 8 : return DICT_H8_CVSS

def get_bus_crit_list(list, bus_crit):
    bus_crit = bus_crit.lower()
    if   bus_crit == 'low' : return list[0]
    elif bus_crit == 'medium' : return list[1]
    elif bus_crit == 'high' or bus_crit == 'high by default' : return list[2]

def calculate_risk(cvss, bus_crit,available_exploit, creds, ip_is_private):
    if cvss == '': cvss = 0
    list = get_cvss_list(cvss)
    list = get_bus_crit_list(list, bus_crit)
    score = 0
    if available_exploit: score += list[1]
    else: score += list[0]
    print(score)
    score += set_access_score(creds, ip_is_private)
    return score

#for i in range(1,10):
#    print(calculate_risk(i, 'high by default', True, False, False))

#tests
#print(True, True, set_access_score(True, True), 1)
#print(True, False, set_access_score(True, False), 3)
#print(False, True, set_access_score(True, False),3)
#print(False, False, set_access_score(False, False), 5)



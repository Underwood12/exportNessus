import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog

rt = tk.Tk()
rt.withdraw()

f = open(filedialog.askopenfilename())

csv.register_dialect('colons', delimiter=';')

reader = csv.reader(f, dialect='colons')
wb = Workbook()
dest_filename = "endfile.xlsx"

ws = wb.worksheets[0]
ws.title = "server criticality"

for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws['%s%s'%(column_letter, (row_index + 1))].value = cell



ws.auto_filter.ref = "A1:" + chr(ws.max_column + 64) + str(ws.max_row)
wb.save(filename = dest_filename)


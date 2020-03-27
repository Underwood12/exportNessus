import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import filedialog


rt = tk.Tk()
rt.withdraw()


def csvs_to_excel(csvs):
    wb = Workbook()
    csv.register_dialect('colons', delimiter=';')
    dest_filename = file_save()
    for file in csvs:
        tmp_ws = wb.create_sheet(file,csvs.index(file))
        with open(file) as f:
            reader = csv.reader(f,delimiter=';')
            for row in reader:
                tmp_ws.append(row)
        tmp_ws.auto_filter.ref = "A1:" + chr(tmp_ws.max_column + 64) + str(tmp_ws.max_row)
    wb.save(filename = dest_filename)


def file_save():
    return filedialog.asksaveasfilename(defaultextension=".xlxs")
